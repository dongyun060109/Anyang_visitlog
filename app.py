import os
import sqlite3
from datetime import date, datetime
import time
import socket

import gradio as gr
import pandas as pd
from fastapi import FastAPI

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ================= 기본 설정(제목 유지) =================
VISITOR_TITLE = "안양 청년1번가 방명록"
ADMIN_TITLE   = "안양 청년1번가 방명록 (관리자 페이지)"

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "1234")

# ================= DB 저장 위치(중요) =================
# Render에서 Persistent Disk를 /var/data 로 마운트할 거야.
DATA_DIR = os.getenv("DATA_DIR", "/var/data")
os.makedirs(DATA_DIR, exist_ok=True)

# ✅ 안양 전용 DB (영구 디스크에 저장)
DB_PATH = os.path.join(DATA_DIR, "visitlog_anyang.db")


GENDER_OPTIONS = ["여성", "남성", "기타"]
AGE_OPTIONS = ["만19~24세", "만25~29세", "만30~34세", "만35~39세"]
RESIDENCE_OPTIONS = ["안양시 동안구", "안양시 만안구", "안양시 비거주(안양활동 청년)", "기타"]
VISIT_TYPE_OPTIONS = ["첫방문", "재방문(2회 이상)"]

PURPOSE_OPTIONS = [
    "공간 프로그램 참여",
    "공부 및 개인작업",
    "미팅 및 워크숍",
    "공용PC, 프린터",
    "간단한 식사 공간",
    "청년 공간이 궁금해서",
    "기타",
]


# ================= DB (잠금에 강한 설정) =================
def get_conn():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn

def init_db():
    conn = get_conn()
    conn.execute("""
    CREATE TABLE IF NOT EXISTS visits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT,
        visit_date TEXT,
        gender TEXT,
        age_group TEXT,
        residence TEXT,
        purpose TEXT,
        visit_type TEXT
    )
    """)
    conn.commit()
    conn.close()

def insert_visit(visit_date, gender, age, residence, purpose, visit_type):
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO visits (created_at, visit_date, gender, age_group, residence, purpose, visit_type)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (datetime.now().isoformat(timespec="seconds"), visit_date, gender, age, residence, purpose, visit_type)
    )
    conn.commit()
    conn.close()

def load_visits(start, end):
    conn = get_conn()
    df = pd.read_sql_query(
        "SELECT * FROM visits WHERE visit_date BETWEEN ? AND ? ORDER BY id ASC",
        conn, params=(start, end)
    )
    conn.close()
    return df

def get_visit_by_id(record_id: int):
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM visits WHERE id = ?", conn, params=(record_id,))
    conn.close()
    return df

def delete_visit(record_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM visits WHERE id = ?", (record_id,))
    conn.commit()
    deleted = cur.rowcount
    conn.close()
    return deleted

def update_visit(record_id: int, visit_date, gender, age, residence, purpose, visit_type):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE visits
        SET visit_date=?, gender=?, age_group=?, residence=?, purpose=?, visit_type=?
        WHERE id=?
    """, (visit_date, gender, age, residence, purpose, visit_type, record_id))
    conn.commit()
    updated = cur.rowcount
    conn.close()
    return updated

def reset_all_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM visits;")
    cur.execute("DELETE FROM sqlite_sequence WHERE name='visits';")
    conn.commit()
    conn.close()


# ================= 메시지(정상 저장: 아무것도 표시 안 함 / 오류만 표시) =================
def warn_box(text: str) -> str:
    return f"""
    <div class="notice notice-warn">
      <div class="notice-title">⚠️ 입력이 필요해요</div>
      <div class="notice-body">{text}</div>
    </div>
    """

def err_box(text: str) -> str:
    return f"""
    <div class="notice notice-err">
      <div class="notice-title">❌ 오류</div>
      <div class="notice-body">{text}</div>
    </div>
    """


# ================= 방문객 저장(정상 저장이면 msg="") =================
def visitor_submit(gender, age, residence, purposes, visit_type, other):
    purposes = list(purposes) if purposes else []
    other = (other or "").strip()

    if gender is None:
        return warn_box("성별을 선택해 주세요."), gender, age, residence, purposes, visit_type, other
    if age is None:
        return warn_box("나이를 선택해 주세요."), gender, age, residence, purposes, visit_type, other
    if residence is None:
        return warn_box("거주지를 선택해 주세요."), gender, age, residence, purposes, visit_type, other
    if visit_type is None:
        return warn_box("방문 횟수를 선택해 주세요."), gender, age, residence, purposes, visit_type, other
    if not purposes:
        return warn_box("방문 목적을 1개 이상 선택해 주세요."), gender, age, residence, purposes, visit_type, other

    plist = [p for p in purposes if p != "기타"]
    if "기타" in purposes:
        plist.append(f"기타:{other}" if other else "기타")

    try:
        for i in range(3):
            try:
                insert_visit(
                    date.today().isoformat(),
                    gender, age, residence,
                    ", ".join(plist),
                    visit_type
                )
                break
            except sqlite3.OperationalError as e:
                if "locked" in str(e).lower() and i < 2:
                    time.sleep(0.4)
                    continue
                raise
    except Exception as e:
        return err_box(f"저장에 실패했어요: {e}"), gender, age, residence, purposes, visit_type, other

    # ✅ 정상 저장이면 메시지 안 뜸
    return "", None, None, None, [], None, ""


# ================= 통계(비율) =================
def calc_ratio(df, col, label):
    if df is None or df.empty:
        return pd.DataFrame(columns=[label, "count", "percent"])
    s = df[col].fillna("").astype(str)
    c = s.value_counts(dropna=False).reset_index()
    c.columns = [label, "count"]
    total = c["count"].sum()
    c["percent"] = (c["count"] / total * 100).round(1) if total else 0
    return c

def calc_purpose_ratio(df):
    if df is None or df.empty:
        return pd.DataFrame(columns=["방문 목적", "count", "percent"])
    temp = df["purpose"].fillna("").astype(str).str.split(", ").explode()
    temp = temp[temp != ""]
    c = temp.value_counts().reset_index()
    c.columns = ["방문 목적", "count"]
    total = c["count"].sum()
    c["percent"] = (c["count"] / total * 100).round(1) if total else 0
    return c


# ================= 일별 방문자 수 + 평균(일요일 자동 제외) =================
def make_daily_counts(df: pd.DataFrame, start: str, end: str):
    try:
        start_dt = datetime.strptime(start, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end, "%Y-%m-%d").date()
    except:
        empty = pd.DataFrame(columns=["날짜", "방문자 수"])
        return empty, "⚠️ 시작일/종료일 형식이 YYYY-MM-DD인지 확인해줘."

    if end_dt < start_dt:
        empty = pd.DataFrame(columns=["날짜", "방문자 수"])
        return empty, "⚠️ 종료일이 시작일보다 빠릅니다."

    all_days = pd.date_range(start=start_dt, end=end_dt, freq="D")
    days_no_sun = all_days[all_days.weekday != 6]  # 일요일 제외

    if len(days_no_sun) == 0:
        empty = pd.DataFrame(columns=["날짜", "방문자 수"])
        msg = f"📌 선택 기간: **{start} ~ {end}**  \n- (일요일 제외) 계산할 날짜가 없습니다."
        return empty, msg

    if df is None or df.empty:
        daily = pd.DataFrame({"날짜": [d.strftime("%Y-%m-%d") for d in days_no_sun], "방문자 수": [0]*len(days_no_sun)})
        num_days = len(days_no_sun)
        total = 0
        avg = 0.0
    else:
        vc = df["visit_date"].astype(str).value_counts()
        daily = pd.DataFrame({"날짜": [d.strftime("%Y-%m-%d") for d in days_no_sun]})
        daily["방문자 수"] = daily["날짜"].map(vc).fillna(0).astype(int)

        num_days = len(days_no_sun)
        total = int(daily["방문자 수"].sum())
        avg = (total / num_days) if num_days else 0.0

    excluded_sundays = int((all_days.weekday == 6).sum())
    msg = f"📌 선택 기간: **{start} ~ {end}**  \n" \
          f"- 제외된 일요일: **{excluded_sundays}일**  \n" \
          f"- 계산일수(일요일 제외): **{num_days}일**  \n" \
          f"- 총 방문(건): **{total}**  \n" \
          f"- 하루 평균 방문자 수(일요일 제외): **{avg:.2f}명/일**"
    return daily, msg


# ================= 엑셀(xlsx) 체크표(병합 헤더) =================
def purpose_to_flags(purpose_str: str):
    s = (purpose_str or "").strip()
    items = [x.strip() for x in s.split(",") if x.strip()]
    flags = {p: 0 for p in PURPOSE_OPTIONS}
    for it in items:
        if it.startswith("기타:") or it == "기타":
            flags["기타"] = 1
        else:
            if it in flags:
                flags[it] = 1
    return flags

def build_checksheet_matrix(df: pd.DataFrame):
    sub_headers = (
        ["연번"] +
        GENDER_OPTIONS +
        AGE_OPTIONS +
        RESIDENCE_OPTIONS +
        PURPOSE_OPTIONS +
        VISIT_TYPE_OPTIONS
    )

    rows = []
    for _, r in df.iterrows():
        rid = int(r["id"])
        gender = str(r["gender"])
        age = str(r["age_group"])
        residence = str(r["residence"])
        visit_type = str(r["visit_type"])
        p_flags = purpose_to_flags(str(r["purpose"]))

        row = [rid]
        row += [1 if k == gender else 0 for k in GENDER_OPTIONS]
        row += [1 if k == age else 0 for k in AGE_OPTIONS]
        row += [1 if k == residence else 0 for k in RESIDENCE_OPTIONS]
        row += [p_flags[k] for k in PURPOSE_OPTIONS]
        row += [1 if k == visit_type else 0 for k in VISIT_TYPE_OPTIONS]
        rows.append(row)

    sums = [0] * len(sub_headers)
    sums[0] = "합계"
    for row in rows:
        for j in range(1, len(sub_headers)):
            sums[j] += int(row[j])

    return sub_headers, sums, rows

def make_checksheet_xlsx(df: pd.DataFrame, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "방문록(체크표)"

    sub_headers, sums, rows = build_checksheet_matrix(df)

    idx = 1
    group_spans = []
    group_spans.append(("구분", idx, idx)); idx += 1
    group_spans.append(("성별", idx, idx + len(GENDER_OPTIONS) - 1)); idx += len(GENDER_OPTIONS)
    group_spans.append(("나이", idx, idx + len(AGE_OPTIONS) - 1)); idx += len(AGE_OPTIONS)
    group_spans.append(("거주지", idx, idx + len(RESIDENCE_OPTIONS) - 1)); idx += len(RESIDENCE_OPTIONS)
    group_spans.append(("방문 목적", idx, idx + len(PURPOSE_OPTIONS) - 1)); idx += len(PURPOSE_OPTIONS)
    group_spans.append(("방문 횟수", idx, idx + len(VISIT_TYPE_OPTIONS) - 1)); idx += len(VISIT_TYPE_OPTIONS)

    header_fill = PatternFill("solid", fgColor="70AD47")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 26

    for title, c1, c2 in group_spans:
        ws.cell(row=1, column=c1).value = title
        ws.cell(row=1, column=c1).fill = header_fill
        ws.cell(row=1, column=c1).font = header_font
        ws.cell(row=1, column=c1).alignment = center
        ws.cell(row=1, column=c1).border = border
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
            for cc in range(c1 + 1, c2 + 1):
                ws.cell(row=1, column=cc).fill = header_fill
                ws.cell(row=1, column=cc).border = border

    for j, h in enumerate(sub_headers, start=1):
        cell = ws.cell(row=2, column=j)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[3].height = 20
    for j, v in enumerate(sums, start=1):
        cell = ws.cell(row=3, column=j)
        cell.value = v
        cell.alignment = center
        cell.border = border
        cell.font = Font(bold=True)

    r0 = 4
    for i, row in enumerate(rows):
        rr = r0 + i
        ws.row_dimensions[rr].height = 18
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=rr, column=j)
            cell.value = v
            cell.alignment = center
            cell.border = border

    for col in range(1, len(sub_headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 8 if col == 1 else 14

    wb.save(out_path)

def export_checksheet_xlsx(ok, start, end):
    if not ok:
        return None
    df = load_visits(start, end)
    out_path = os.path.join(DATA_DIR, "visitlog_checksheet.xlsx")
    make_checksheet_xlsx(df, out_path)
    return out_path


# ================= 관리자 기능 =================
def admin_load_all(ok, s, e):
    if not ok:
        empty = pd.DataFrame()
        empty_daily = pd.DataFrame(columns=["날짜", "방문자 수"])
        return empty, empty, empty, empty, empty, empty, empty, empty_daily, ""
    df = load_visits(s, e)
    daily_df, daily_msg = make_daily_counts(df, s, e)
    return (
        df,
        calc_purpose_ratio(df),
        calc_ratio(df, "gender", "성별"),
        calc_ratio(df, "age_group", "나이"),
        calc_ratio(df, "residence", "거주지"),
        calc_ratio(df, "visit_type", "방문 횟수"),
        df,
        daily_df,
        daily_msg
    )

def admin_fetch_one(ok, record_id):
    if not ok:
        return "❌ 관리자 로그인 필요", "", "여성", "만19~24세", "안양시 동안구", [], "첫방문", ""
    try:
        rid = int(record_id)
    except:
        return "⚠️ ID는 숫자로 입력", "", "여성", "만19~24세", "안양시 동안구", [], "첫방문", ""

    df = get_visit_by_id(rid)
    if df.empty:
        return f"⚠️ ID {rid} 데이터 없음", "", "여성", "만19~24세", "안양시 동안구", [], "첫방문", ""

    r = df.iloc[0]
    visit_date = str(r["visit_date"])
    gender = str(r["gender"])
    age = str(r["age_group"])
    residence = str(r["residence"])
    visit_type = str(r["visit_type"])
    purpose_str = str(r["purpose"])

    purpose_items = [x.strip() for x in purpose_str.split(",") if x.strip()]
    purposes = []
    other_text = ""
    for it in purpose_items:
        if it.startswith("기타:"):
            purposes.append("기타")
            other_text = it.replace("기타:", "", 1).strip()
        elif it == "기타":
            purposes.append("기타")
        elif it in PURPOSE_OPTIONS:
            purposes.append(it)

    return "✅ 불러왔어. 아래에서 수정 후 [수정 저장] 눌러.", visit_date, gender, age, residence, purposes, visit_type, other_text

def admin_update_one(ok, record_id, visit_date, gender, age, residence, purposes, visit_type, other_text):
    if not ok:
        return "❌ 관리자 로그인 필요"
    try:
        rid = int(record_id)
    except:
        return "⚠️ ID는 숫자로 입력"

    visit_date = (visit_date or "").strip()
    try:
        datetime.strptime(visit_date, "%Y-%m-%d")
    except:
        return "⚠️ 방문일은 YYYY-MM-DD 형식"

    purposes = list(purposes) if purposes else []
    other_text = (other_text or "").strip()
    if not purposes:
        return "⚠️ 방문 목적 1개 이상 선택"

    plist = [p for p in purposes if p != "기타"]
    if "기타" in purposes:
        plist.append(f"기타:{other_text}" if other_text else "기타")
    purpose_str = ", ".join(plist)

    updated = update_visit(rid, visit_date, gender, age, residence, purpose_str, visit_type)
    if updated == 0:
        return f"⚠️ ID {rid} 업데이트 실패(없을 수 있음)"
    return f"✅ ID {rid} 수정 저장 완료"

def request_delete(ok, record_id):
    if not ok:
        return "❌ 관리자 로그인 필요", None, gr.update(visible=False)
    try:
        rid = int(record_id)
    except:
        return "⚠️ ID는 숫자로 입력", None, gr.update(visible=False)

    df = get_visit_by_id(rid)
    if df.empty:
        return f"⚠️ ID {rid} 데이터 없음", None, gr.update(visible=False)

    msg = f"⚠️ 정말 삭제할까요? (ID={rid}) 아래 [정말 삭제]를 한 번 더 눌러야 삭제됩니다."
    return msg, rid, gr.update(visible=True)

def confirm_delete(ok, pending_id):
    if not ok:
        return "❌ 관리자 로그인 필요", None, gr.update(visible=False)
    if pending_id is None:
        return "⚠️ 먼저 [삭제]를 눌러 삭제 확인을 받아야 해.", None, gr.update(visible=False)

    deleted = delete_visit(int(pending_id))
    if deleted == 0:
        return f"⚠️ ID {pending_id} 삭제 실패(없을 수 있음)", None, gr.update(visible=False)
    return f"✅ ID {pending_id} 삭제 완료", None, gr.update(visible=False)

def admin_reset_all(ok, pw1, pw2):
    if not ok:
        return "❌ 관리자 로그인 필요"
    pw1 = (pw1 or "").strip()
    pw2 = (pw2 or "").strip()
    if not pw1 or not pw2:
        return "⚠️ 비밀번호를 2칸 모두 입력해줘."
    if pw1 != pw2:
        return "⚠️ 두 비밀번호가 서로 달라."
    if pw1 != ADMIN_PASSWORD:
        return "❌ 비밀번호가 틀렸어."
    reset_all_data()
    return "✅ 전체 데이터 초기화 완료! (연번도 1부터 다시 시작)"


# ================= 방문자 CSS(버튼 크기/비율 그대로) =================
VISITOR_CSS = """
.gradio-container { max-width: 720px !important; margin: 0 auto !important; }

.vlabel{
  font-size: 15px !important;
  font-weight: 700 !important;
  margin: 8px 0 10px 2px !important;
  opacity: 0.95 !important;
}

/* 선택 버튼 크기/비율 유지 */
.gradio-container .gr-radio label,
.gradio-container .gr-checkboxgroup label,
.gradio-container [role="radiogroup"] label,
.gradio-container [role="group"] label{
  display:flex !important;
  align-items:center !important;

  padding: 44px 30px !important;
  margin: 16px 0 !important;
  border-radius: 18px !important;
  min-height: 96px !important;

  font-size: 32px !important;
  line-height: 1.12 !important;
  gap: 24px !important;
}

.gradio-container input[type="radio"],
.gradio-container input[type="checkbox"]{
  width: 38px !important;
  height: 38px !important;
  min-width: 38px !important;
  min-height: 38px !important;
}

/* 저장 버튼 크기/비율 유지 */
#save_btn button,
#save_btn{
  font-size: 30px !important;
  padding: 26px 22px !important;
  border-radius: 18px !important;
  min-height: 84px !important;
}

/* 경고/오류 박스 강조 (정상 저장 시엔 msg가 빈 문자열) */
.notice{
  border-radius: 18px !important;
  padding: 18px 18px !important;
  margin: 14px 0 16px 0 !important;
  border: 3px solid transparent !important;
}

.notice-title{
  font-size: 28px !important;
  font-weight: 900 !important;
  margin-bottom: 8px !important;
  letter-spacing: -0.3px !important;
}

.notice-body{
  font-size: 24px !important;
  font-weight: 800 !important;
  line-height: 1.25 !important;
}

.notice-warn{
  background: rgba(255, 120, 0, 0.18) !important;
  border-color: rgba(255, 120, 0, 0.95) !important;
  box-shadow: 0 0 0 3px rgba(255,120,0,0.22) inset, 0 10px 22px rgba(0,0,0,0.10) !important;
}

.notice-err{
  background: rgba(239, 68, 68, 0.16) !important;
  border-color: rgba(239, 68, 68, 0.95) !important;
  box-shadow: 0 0 0 3px rgba(239,68,68,0.18) inset !important;
}
"""


# ================= Gradio 앱 생성(launch 하지 않음) =================
def build_visitor_app():
    with gr.Blocks(css=VISITOR_CSS, title=VISITOR_TITLE) as visitor_app:
        gr.Markdown(f"# 👤 {VISITOR_TITLE}")
        gr.Markdown("항목을 선택하고 **저장**을 눌러주세요. 저장 후 자동으로 초기화됩니다. *(방문일은 오늘 자동 저장)*")

        msg = gr.HTML("")

        gr.Markdown('<div class="vlabel">성별</div>')
        g = gr.Radio(GENDER_OPTIONS, value=None, show_label=False)

        gr.Markdown('<div class="vlabel">나이</div>')
        a = gr.Radio(AGE_OPTIONS, value=None, show_label=False)

        gr.Markdown('<div class="vlabel">거주지</div>')
        r = gr.Radio(RESIDENCE_OPTIONS, value=None, show_label=False)

        gr.Markdown('<div class="vlabel">방문 목적 (복수 선택)</div>')
        p = gr.CheckboxGroup(PURPOSE_OPTIONS, value=[], show_label=False)

        gr.Markdown('<div class="vlabel">기타 내용(선택)</div>')
        o = gr.Textbox(value="", show_label=False)

        gr.Markdown('<div class="vlabel">방문 횟수</div>')
        v = gr.Radio(VISIT_TYPE_OPTIONS, value=None, show_label=False)

        btn = gr.Button("✅ 저장", variant="primary", elem_id="save_btn")
        btn.click(visitor_submit, [g, a, r, p, v, o], [msg, g, a, r, p, v, o])

    return visitor_app

def build_admin_app():
    with gr.Blocks(title=ADMIN_TITLE) as admin_app:
        gr.Markdown(f"# 🔒 {ADMIN_TITLE}")

        pw = gr.Textbox(label="비밀번호", type="password")
        ok = gr.State(False)
        login_btn = gr.Button("로그인", variant="primary")
        login_msg = gr.Markdown("")

        def do_login(p):
            okv = (p or "").strip() == ADMIN_PASSWORD
            return okv, ("✅ 로그인 성공" if okv else "❌ 비밀번호가 틀렸어")

        login_btn.click(do_login, pw, [ok, login_msg])

        start = gr.Textbox(value=date.today().replace(day=1).isoformat(), label="시작일 (YYYY-MM-DD)")
        end   = gr.Textbox(value=date.today().isoformat(), label="종료일 (YYYY-MM-DD)")
        load_btn = gr.Button("데이터/비율 불러오기", variant="primary")

        table = gr.Dataframe(label="원본 데이터(조회)", interactive=False)

        purpose_ratio   = gr.Dataframe(label="방문 목적 비율(%)", interactive=False)
        gender_ratio    = gr.Dataframe(label="성별 비율(%)", interactive=False)
        age_ratio       = gr.Dataframe(label="나이 비율(%)", interactive=False)
        residence_ratio = gr.Dataframe(label="거주지 비율(%)", interactive=False)
        visit_type_ratio= gr.Dataframe(label="방문 횟수 비율(%)", interactive=False)

        gr.Markdown("## 📅 일별 방문자 수 & 하루 평균 (일요일 자동 제외)")
        daily_table = gr.Dataframe(label="일별 방문자 수(일요일 제외, 0명인 날 포함)", interactive=False)
        daily_avg_md = gr.Markdown("")

        export_btn  = gr.Button("✅ 엑셀(xlsx) 체크표 다운로드(병합 헤더)", variant="secondary")
        export_file = gr.File(label="다운로드 파일")

        gr.Markdown("## ✏️ 원본 데이터 수정/삭제 (ID로 작업)")
        record_id = gr.Textbox(label="수정/삭제할 ID(연번)", placeholder="예: 3487")
        fetch_btn = gr.Button("ID로 불러오기")
        edit_status = gr.Markdown("")

        edit_date = gr.Textbox(label="방문일 (YYYY-MM-DD)")
        edit_gender = gr.Radio(GENDER_OPTIONS, label="성별", value="여성")
        edit_age    = gr.Radio(AGE_OPTIONS, label="나이", value="만19~24세")
        edit_res    = gr.Radio(RESIDENCE_OPTIONS, label="거주지", value="안양시 동안구")
        edit_purposes = gr.CheckboxGroup(PURPOSE_OPTIONS, label="방문 목적(복수 선택)")
        edit_visit_type = gr.Radio(VISIT_TYPE_OPTIONS, label="방문 횟수", value="첫방문")
        edit_other = gr.Textbox(label="기타 내용(선택)")

        update_btn = gr.Button("✅ 수정 저장", variant="primary")

        pending_delete_id = gr.State(None)
        delete_btn = gr.Button("🗑️ 삭제", variant="stop")
        confirm_delete_btn = gr.Button("⚠️ 정말 삭제", variant="stop", visible=False)
        action_msg = gr.Markdown("")

        gr.Markdown("## 🔥 전체 데이터 초기화 (비밀번호 2중 확인)")
        reset_pw1 = gr.Textbox(label="비밀번호 입력(1)", type="password")
        reset_pw2 = gr.Textbox(label="비밀번호 입력(2)", type="password")
        reset_btn = gr.Button("⚠️ 전체 데이터 초기화 실행", variant="stop")
        reset_msg = gr.Markdown("")

        load_btn.click(
            admin_load_all,
            [ok, start, end],
            [table, purpose_ratio, gender_ratio, age_ratio, residence_ratio, visit_type_ratio,
             table, daily_table, daily_avg_md]
        )
        export_btn.click(export_checksheet_xlsx, [ok, start, end], [export_file])

        fetch_btn.click(
            admin_fetch_one,
            [ok, record_id],
            [edit_status, edit_date, edit_gender, edit_age, edit_res, edit_purposes, edit_visit_type, edit_other]
        )
        update_btn.click(
            admin_update_one,
            [ok, record_id, edit_date, edit_gender, edit_age, edit_res, edit_purposes, edit_visit_type, edit_other],
            [action_msg]
        )
        delete_btn.click(
            request_delete,
            [ok, record_id],
            [action_msg, pending_delete_id, confirm_delete_btn]
        )
        confirm_delete_btn.click(
            confirm_delete,
            [ok, pending_delete_id],
            [action_msg, pending_delete_id, confirm_delete_btn]
        )
        reset_btn.click(admin_reset_all, [ok, reset_pw1, reset_pw2], [reset_msg])

    return admin_app


# ================= FastAPI에 Gradio 2개 URL로 붙이기 =================
init_db()
visitor_app = build_visitor_app()
admin_app   = build_admin_app()

app = FastAPI()
app = gr.mount_gradio_app(app, visitor_app, path="/")
app = gr.mount_gradio_app(app, admin_app,   path="/admin")

